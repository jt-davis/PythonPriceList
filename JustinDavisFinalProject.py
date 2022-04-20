# INF360 - Programming in Python
# Justin Davis
# Final Project


'''

This project allows the user to manipulate a dictionary of products and
prices for a specific business. In this case it's a business that builds
wood decor and furniture out of reclaimed pallets. The openpyxl module
must be downloaded and installed in order for this project to run.

First, the program loads an initial dictionary of the business' current
products and their prices. Then, it brings up a menu in which the user selects
an option by typing it in. The options are different forms of manipulating the
data in the dictionary (add items, delete items, sort the dictionary by name,
sort the dictionary by price, print the dictionary in a concise format, and
export the dictionary to an excel file). Once done, the user can choose to exit
the program.

'''

import sys

import re

import logging
logging.basicConfig(filename='logfilename.log', level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

# This program needs the openpyxl module to run
try:
    logging.debug('Importing openpyxl module')
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
except:
    logging.critical('Program about to crash. openpyxl not found.')
    print('openpyxl was not found')

logging.debug('openpyxl imported successfully')
logging.debug('Start of program')


priceSheet = {'Door Shelf (Oak)': 55.00,
              'Serving Tray (Oak)': 40.00,
              'Jewelry Hanger (Pine)': 35.00,
              'Orange Pumpkins set of 3 (Pine)': 40.00,
              'Orange Pumpkins set of 3 (Oak)': 40.00,
              'Clock (Oak and Pine)': 60.00,
              'End Table (Oak)': 60.00,
              'Coat Rack (Oak and Pine)': 50.00,
              'Toilet Paper Holder (Pine)': 45.00,
              'Wine Rack Small (Oak)': 45.00,
              'Spice and Utensil Holder (Pine)': 30.00,
              'Three Tier Wall Shelf (Oak and Pine)': 40.00,
              'Coffee Table (Oak and Pine)': 100.00,
              'Small Shelf/Towel Hanger (Oak)': 40.00,
              'Cross (Oak)': 30.00,
              'Night Stand (Oak)': 80.00,
              'DVD Holders (Pine)': 40.00,
              'Small Floating Shelf (Oak and Pine)': 45.00,
              'Large Floating Shelf (Oak and Pine)': 75.00,
              'Small 3D Shelf (Oak)': 45.00,
              'Large 3D Shelf (Oak and Pine)': 60.00,
              'Entryway Table (Oak and Pine)': 85.00,
              'Small Custom Word Wall Sign (Oak and Pine)': 50.00,
              'Large Three Tier Shelf (Oak and Pine)': 55.00,
              'Corner Coffee Table (Oak)': 85.00}

def printPrice(sheet):
    logging.debug('Starting printPrice function')
    logging.debug('Creating empty dictionary called lengthList')
    lengthList = []
    for k in sheet.keys():
        # Adds the length of each key to lengthList
        logging.debug('The length of the key is: ' + str(len(k)))
        lengthList.append((len(k)))
        logging.debug('List updated with length')
    logging.debug('Max length is: ' + str(max(lengthList)))
    logging.debug('Printing dictionary of item name and price')         
    for k, v in sheet.items():
        # Gets the longest key and adds eight spaces
        print(k.ljust(max(lengthList) + 8, '.') + '$'.rjust(2) + str(format(v, '.2f')))
    logging.debug('End of printPrice function')   
    print('')

# Creates a function for adding an item to the dictionary
def addItem(dictionary):
    logging.debug('Starting addItem function')
    while True:
        print('What is the name of the new product? Please type it in this format: Item Name (wood type)')
        productName = input()
        # Creates a regex that searches for any number of any characters, then a space, then a '(',
        # then any number of letters only, then a ')'
        logging.debug('Creating productRegex')
        productRegex = re.compile(r'^\b.+\B[(]\b[a-zA-Z ]+\b[)]$')          
        while True:
            # Applies the regex to the user's input
            logging.debug('Applying productRegex to user\'s input under the variable mo')
            mo = productRegex.search(productName)
            # If the user entered the input in the wrong format it will ask them again until they do
            if mo == None:
                logging.debug('mo is None')
                print('')
                print('Please type the name of the new product in this format: Item Name (wood type)')      
                productName = input()                                                                       
            else:
                break

        print('')     
            
        print('What is the price of the new product? Please type it in this format: 00.00')
        productPrice = input()
        # Creates a regex that searches for one or more digits,
        # then a '.', then exactly two digits
        logging.debug('Creating priceRegex')
        priceRegex = re.compile(r'^\d+[.]\d{2}$')               
        while True:
            # Applies the regex to the user's input
            logging.debug('Applying priceRegex to user\'s input under the variable mo')
            mo = priceRegex.search(productPrice)
            # If the user entered the input in the wrong format it will ask them again until they do
            if mo == None:
                logging.debug('mo is None')
                print('')
                print('Please type the price in this format: 00.00')
                productPrice = input()
            else:
                 break
        # Adds the user's inputs into the original dictionary
        logging.debug('Adding the user\'s inputs into the original dictionary')
        dictionary[productName] = float(productPrice)      
        logging.debug('Dictionary updated')

        print('')

        print('Would you like to print the updated list?')
        printListAnswer = input()
        # Continually asks the user until they answer yes or no exactly, case insensitive
        while True:                                 
            if printListAnswer.upper() == 'YES':
                 print('')
                 logging.debug('Printing updated dictionary')
                 printPrice(dictionary)
                 break
            elif printListAnswer.upper() == 'NO':
                 print('Okay.')
                 break
            else:
                 print('Please type yes or no.')
                 printListAnswer = input()
        break

    logging.debug('End of addItem function')

def deleteItem(dictionary):
    logging.debug('Starting deleteItem function')
    # Prints the original priceSheet for easy reference
    logging.debug('Printing original dictionary for easy reference')
    printPrice(dictionary)                          
    print('Type the item you would like to delete exactly as it is.')
    deleteItem = input()
    while True:
        # Checks to see if the user's input is a key in the dictionary
        # If it is, it will delete the entire item from the dictionary
        logging.debug('Checking if user\'s input is in the dictionary')
        if deleteItem in dictionary.keys():
            logging.debug('User\'s input is in the dictionary')
            logging.debug('Deleting user\'s input from the dictionary')
            del dictionary[deleteItem]
            logging.debug('Item deleted')
            break
        else:
            # If the user's input is not a key in the dictionary, the user is
            # asked again until it matches something in the dictionary
            logging.debug('User\'s input is not in the dictionary')
            print('Please enter the item you want to delete exactly.')      
            deleteItem = input()                                            
    print('')
    print('Item deleted. Would you like to print the updated list?')        
    printAnswer = input()
    print('')
    while True:
        # Continually asks the user until they answer yes or no exactly, case insensitive
        if printAnswer.upper() == 'YES':
            logging.debug('Printing the updated dictionary')
            printPrice(dictionary)
            break
        elif printAnswer.upper() == 'NO':
            print('Okay.')
            break
        else:
            print('Please type yes or no.')
            printAnswer = input()
    logging.debug('End of deleteItem function')
            
def sortList(dictionary):
    logging.debug('Starting sortList function')
    # Creates a list out of the keys of the original dictionary
    logging.debug('Creating a new list out of the keys from the original dictionary under the variable listOfKeys')
    listOfKeys = list(dictionary.keys())
    # Sorts that new list alphabetically
    logging.debug('Sorting listOfKeys')
    listOfKeys.sort(key=str.lower)
    # Creates an empty dictionary
    logging.debug('Creating and empty dictionary under the variable sortedSheet')
    sortedSheet = {}
    # Takes each list item in the new list, makes it a key in the new dictionary, then grabs the
    # value of the key from the old dictionary and makes it the value of the key in the new dictionary
    for key in listOfKeys:
        logging.debug('Adding the keys to sortedSheet, then adding its value from the dictionary')
        sortedSheet[key] = dictionary[key]
    # Runs the new dictionary through the printPrice function
    logging.debug('Printing sortedSheet')
    printPrice(sortedSheet)
    logging.debug('End of sortList function')

def sortByPrice(dictionary):
    logging.debug('Starting sortByPrice function')
    # Creates an empty dictionary
    logging.debug('Creating empty dictionary under the variable sortedSheetByPrice')
    sortedSheetByPrice = {}
    # Creates a list full of the values from the dictionary passed to it
    logging.debug('Creating a list of the values in the dictionary under the variable listOfValues')
    listOfValues = list(dictionary.values())
    # Sorts the list in reverse
    logging.debug('Sorting listOfValues in reverse order')
    listOfValues.sort(reverse=True)
    # Goes through every value in the list, then goes through the items in the dictionary passed to it
    # and checks if the value in that dictionary equals the value in the list of values
    # If it does, it adds the key from the passed dictionary as a key in the new dictionary with the new
    # value as the value that matched it in the original dictionary
    logging.debug('Adding keys and values to sortedSheetByPrice, while keeping them in sorted order')
    for value in listOfValues:
        for k, v in dictionary.items():
            if v == value:
                sortedSheetByPrice[k] = v
    # Runs the new dictionary, which is just a sorted version of the old dictionary, through the print function
    logging.debug('Printing sortedSheetByPrice')
    printPrice(sortedSheetByPrice)
    logging.debug('End of sortByPrice function')

def exportToExcel(dictionary):
    logging.debug('Starting exportToExcel function')
    logging.debug('Creating workbook object')
    wb = openpyxl.Workbook()
    sheet = wb.active
    logging.debug('Naming sheet Price List')
    sheet.title = 'Price List'
    # Sets header info
    logging.debug('Formatting header cells')
    for i in range (1,3):
        sheet.cell(row = 1, column = i).font = Font(bold = 'True', size = '20')
        sheet.cell(row = 1, column = i).alignment = Alignment(horizontal = 'center')
    sheet.cell(row = 1, column = 1).value = 'Item'
    sheet.cell(row = 1, column = 2).value = 'Price'
    logging.debug('Headers formatted')
    # Adds keys and values of dictionary to Excel cells
    logging.debug('Writing dictionary keys and values to cells in the Excel file')
    row = 2
    for k, v in dictionary.items():
        sheet.cell(row = row, column = 1).value = k
        sheet.cell(row = row, column = 2).value = v
        row += 1
    logging.debug('Excel file succesfully created with keys and values of dictionary')
    # Adjusts first column width
    logging.debug('Adjusting width of column A')
    sheet.column_dimensions['A'].width = 35
    # Saves Excel file, errors if it didn't save for some reason
    try:
        logging.debug('Saving Excel file under the name Price_List.xlsx')
        wb.save('Price_List.xlsx')
    except:
        logging.critical('Excel file could not save.')
        print('Error: Excel file could not save.')
    logging.debug('Excel file saved sucessfully')
    print('Excel file created and saved under the name Price_List.xlsx')
                

while True:
    # Creates a menu that runs at the start of the program and after every function is finished
    logging.debug('Printing menu')
    print('Menu'.center(31, '-'))               
    print('Add to list'.ljust(15) + 'Delete from list'.rjust(0))
    print('Sort by name'.ljust(15) + 'Sort by price'.rjust(0))
    print('Print list'.ljust(15) + 'Exit program'.rjust(0))
    print('Export to Excel'.ljust(15))
    print(''.center(31, '-'))
    menuAnswer = input()
    print('')
    while True:
        # Continually asks the user until they answer add, delete, print, sort, or exit
        if menuAnswer.upper() == 'ADD TO LIST':
            addItem(priceSheet)
            break
        elif menuAnswer.upper() == 'DELETE FROM LIST':
            deleteItem(priceSheet)
            break
        elif menuAnswer.upper() == 'PRINT LIST':
            printPrice(priceSheet)
            break
        elif menuAnswer.upper() == 'SORT BY NAME':
            sortList(priceSheet)
            break
        elif menuAnswer.upper() == 'SORT BY PRICE':         
            sortByPrice(priceSheet)
            break
        elif menuAnswer.upper() == 'EXPORT TO EXCEL':
            exportToExcel(priceSheet)
            break
        elif menuAnswer.upper() == 'EXIT PROGRAM':
            logging.debug('End of program')
            sys.exit()
        else:
            print('Please enter a menu option exactly as it appears.')
            menuAnswer = input()
